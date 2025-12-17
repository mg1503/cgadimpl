#!/usr/bin/env python3
"""Script to add team task breakdown sheet to plan_v2.xlsx"""

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load the workbook
wb = load_workbook('plan_v2.xlsx')

# Create a new sheet
ws = wb.create_sheet("Team Tasks")

# Define the data structure
teams_data = [
    {
        'team': 'Team 1',
        'members': [
            {
                'name': 'kurra',
                'tasks': [
                    {'task': 'debugging present checkpointing', 'days': 4},
                    {'task': 'adaptive checkpointing impl', 'days': 3},
                    {'task': 'chens algo impl', 'days': 5}
                ]
            },
            {
                'name': 'surendra',
                'tasks': [
                    {'task': 'read&understand existing checkpointing in general', 'days': 4},
                    {'task': 'understand checkpointing of framework', 'days': 3},
                    {'task': 'uniform checkpoint impl', 'days': 3},
                    {'task': 'forward recompute in checkpointing impl', 'days': 2}
                ]
            }
        ]
    },
    {
        'team': 'Team 2',
        'members': [
            {
                'name': 'jenifa',
                'tasks': [
                    {'task': 'thread pool impl and basic worker loop', 'days': 1.5},
                    {'task': 'integrate readyqueue ith worker and shutdown logic', 'days': 1.5},
                    {'task': 'merge & integrate to backward()', 'days': 1},
                    {'task': 'refactor half the functions adding this functionalit to nodeops.', 'days': 2},
                    {'task': 'testing', 'days': 2},
                    {'task': 'brainstrom on edge cases', 'days': 1},
                    {'task': 'implement the edge cases', 'days': 2},
                    {'task': 'learn cuda_Stream', 'days': 3}
                ]
            },
            {
                'name': 'grishma',
                'tasks': [
                    {'task': 'implement dependency counter map and initialization', 'days': 1.5},
                    {'task': 'implement counter decrement logic and parent readying', 'days': 1.5},
                    {'task': 'merge and intigrate backward', 'days': 1},
                    {'task': 'refactoring nodeops', 'days': 2},
                    {'task': 'testing', 'days': 2},
                    {'task': 'brainstrom on edge cases', 'days': 1},
                    {'task': 'implement the edge cases', 'days': 2},
                    {'task': 'learn cuda_Stream', 'days': 3}
                ]
            }
        ]
    },
    {
        'team': 'Team 3',
        'members': [
            {
                'name': 'rishi',
                'tasks': [
                    {'task': 'general disection and outline', 'days': 3},
                    {'task': 'CUDA mem pool(disection)', 'days': 3},
                    {'task': 'with stream cudagraph(toy implementation based of our fraamework)', 'days': 4},
                    {'task': 'testing', 'days': 2}
                ]
            },
            {
                'name': 'jona',
                'tasks': [
                    {'task': 'general disection and outline', 'days': 3},
                    {'task': 'graph capture and executor', 'days': 3},
                    {'task': 'w/o cudastrem cudagraph(toy implementation based of our fraamework)', 'days': 4},
                    {'task': 'testing', 'days': 2}
                ]
            }
        ]
    }
]

# Prepare data for the chart and table
row = 1

# Title
ws.merge_cells('A1:E1')
ws['A1'] = 'Team Task Breakdown - Timeline Visualization'
ws['A1'].font = Font(size=16, bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Headers for visualization
row = 3
ws['A3'] = 'Team'
ws['B3'] = 'Member'
ws['C3'] = 'Task'
ws['D3'] = 'Days'
ws['E3'] = 'Cumulative Days'

# Style headers
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
for col in ['A', 'B', 'C', 'D', 'E']:
    ws[f'{col}3'].fill = header_fill
    ws[f'{col}3'].font = header_font
    ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')

# Populate data
row = 4
chart_data_rows = []
for team in teams_data:
    for member in team['members']:
        cumulative = 0
        for task in member['tasks']:
            ws[f'A{row}'] = team['team']
            ws[f'B{row}'] = member['name']
            ws[f'C{row}'] = task['task']
            ws[f'D{row}'] = task['days']
            cumulative += task['days']
            ws[f'E{row}'] = cumulative
            
            # Center align
            for col in ['A', 'B', 'D', 'E']:
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            chart_data_rows.append(row)
            row += 1

# Adjust column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 18

# Create bar chart
chart = BarChart()
chart.type = "bar"
chart.style = 11
chart.title = "Task Duration by Person"
chart.x_axis.title = "Days"
chart.y_axis.title = "Tasks"

# Add data to chart
data = Reference(ws, min_col=4, min_row=3, max_row=row-1)
categories = Reference(ws, min_col=3, min_row=4, max_row=row-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Chart appearance
chart.height = 20
chart.width = 25

# Position chart
chart_start_row = row + 2
ws.add_chart(chart, f'A{chart_start_row}')

# Add checkbox-based todo list below chart
todo_start_row = chart_start_row + 35

ws.merge_cells(f'A{todo_start_row}:E{todo_start_row}')
ws[f'A{todo_start_row}'] = 'Task Checklist'
ws[f'A{todo_start_row}'].font = Font(size=14, bold=True)
ws[f'A{todo_start_row}'].alignment = Alignment(horizontal='center', vertical='center')

# Headers for todo
todo_start_row += 2
ws[f'A{todo_start_row}'] = '☐'
ws[f'B{todo_start_row}'] = 'Team'
ws[f'C{todo_start_row}'] = 'Member'
ws[f'D{todo_start_row}'] = 'Task'
ws[f'E{todo_start_row}'] = 'Days'

# Style todo headers
for col in ['A', 'B', 'C', 'D', 'E']:
    ws[f'{col}{todo_start_row}'].fill = header_fill
    ws[f'{col}{todo_start_row}'].font = header_font
    ws[f'{col}{todo_start_row}'].alignment = Alignment(horizontal='center', vertical='center')

# Populate todo list
todo_row = todo_start_row + 1
for team in teams_data:
    for member in team['members']:
        for task in member['tasks']:
            ws[f'A{todo_row}'] = '☐'
            ws[f'B{todo_row}'] = team['team']
            ws[f'C{todo_row}'] = member['name']
            ws[f'D{todo_row}'] = task['task']
            ws[f'E{todo_row}'] = task['days']
            
            # Center align
            for col in ['A', 'B', 'E']:
                ws[f'{col}{todo_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{todo_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'D{todo_row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            # Add light border
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{todo_row}'].border = thin_border
            
            todo_row += 1

# Save the workbook
wb.save('plan_v2.xlsx')
print("Successfully added 'Team Tasks' sheet to plan_v2.xlsx")
